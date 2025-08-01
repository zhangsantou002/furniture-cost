from openpyxl import Workbook

def export_to_excel(data, filename):
    """
    将数据导出为 Excel 文件。
    data: list of dict
    filename: 导出的文件名
    """
    wb = Workbook()
    ws = wb.active
    
    if not data:
        wb.save(filename)
        return
    
    # 写入表头
    headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入数据
    for row, item in enumerate(data, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=item.get(header, ''))
    
    wb.save(filename)